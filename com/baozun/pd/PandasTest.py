import pandas as pd
import os
import openpyxl

def testSeries():
    data = [1, 2, 3, 7, 'ab']
    index = ['a', 'b', 'c', 'd', 'f']
    s = pd.Series(data, index=index)
    print(f"index={s.index}")
    print(f"values={s.values}")
    print(f"dtype={s.dtype}")
    print(f"head={s.head(2)}")
    print('------------')
    print(f"s == {s}")
    newS = s.map(lambda x: x * 2)
    print(f"ns == {newS}")

def testExcel():
    path = 'D:\\file\\Downloads\\test'
    fileName = 'output.xlsx'
    newPath = os.path.join(path, fileName)
    excel_data = pd.read_excel(newPath)
    # 将DataFrame转换为JSON
    # json_data = excel.to_json(orient='records', indent=2)
    # print(json_data)
    # excel.to_json(os.path.join(path, 'output.json'), orient='records', indent=2)

def openpyxlRead():
    path = 'D:\\file\\Downloads\\test'
    fileName = 'output.xlsx'
    newPath = os.path.join(path, fileName)
    try:
        workbook = openpyxl.load_workbook(newPath)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"sheet名称={sheet_name}")
            print(f"行数={sheet.max_row}")
            print(f"列数={sheet.max_column}")
            # 读取第一行
            first_row = [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)]
            print(f"第一行数据: {first_row}")
            # 读取第一列
            first_col = [sheet.cell(row=r, column=1).value for r in range(1, sheet.max_row + 1)]
            print(f"第一列数据: {first_col}")

            # 读取第一个单元格
            first_cell = sheet.cell(row=1, column=1).value
            print(f"第一个单元格值: {first_cell}")

            cellValue = sheet['A:B']
            print(f"A:B = {cellValue}")

            iterRows = sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=4)
            print(f'iterRows = {iterRows}')

    except FileNotFoundError as ex:
        print(f"文件 {newPath} 不存在,ex {ex}")
    except Exception as e:
        print(f"发生错误：{e}")






if __name__ == '__main__':
    # testSeries()
    # testExcel()
    openpyxlRead()